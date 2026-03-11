from __future__ import annotations

from collections.abc import Generator
from pathlib import Path
from typing import Any, Optional
from zipfile import ZipFile, is_zipfile
import logging
import tempfile
import os
import struct

import olefile
import xlrd
from openpyxl import load_workbook

from dify_plugin.entities import I18nObject
from dify_plugin.entities.tool import ToolInvokeMessage, ToolParameter
from dify_plugin import Tool
from dify_plugin.file.file import File


class ExcelExtractorTool(Tool):
    """
    Extracts textual cell content and embedded images from Excel spreadsheets.
    Supports both OOXML (.xlsx) and legacy Binary Interchange File Format (.xls).
    """

    _CONTINUE_RECORD = 0x003C
    _MSODRAWINGGROUP_RECORD = 0x00EB
    _BSE_RECORD_TYPE = 0xF007
    _BLIP_RECORD_TYPES = {
        0xF018,
        0xF019,
        0xF01A,
        0xF01B,
        0xF01C,
        0xF01D,
        0xF01E,
        0xF01F,
        0xF020,
        0xF021,
        0xF029,
        0xF02A,
        0xF02B,
        0xF02C,
    }

    def _invoke(
        self,
        tool_parameters: dict[str, Any],
        user_id: Optional[str] = None,
        conversation_id: Optional[str] = None,
        app_id: Optional[str] = None,
        message_id: Optional[str] = None,
    ) -> Generator[ToolInvokeMessage, None, None]:
        logger = logging.getLogger(__name__)

        try:
            excel_content = tool_parameters.get("excel_content")
            if not isinstance(excel_content, File):
                raise ValueError("Invalid Excel content format. Expected File object.")

            original_filename = excel_content.filename or "workbook"
            suffix = self._determine_suffix(original_filename)

            logger.info(
                "[excel_extractor] start extraction",
                extra={
                    "original_filename": original_filename,
                    "suffix": suffix,
                    "mime_type": excel_content.mime_type,
                },
            )

            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
                temp_file.write(excel_content.blob)
                temp_file_path = temp_file.name

            try:
                is_modern_excel = self._is_ooxml_excel(temp_file_path)
                logger.info(
                    "[excel_extractor] detected excel format",
                    extra={"is_modern_excel": is_modern_excel},
                )

                if is_modern_excel:
                    text_content = self._extract_text_xlsx(temp_file_path)
                    images = list(self._extract_images_xlsx(temp_file_path))
                else:
                    text_content = self._extract_text_xls(temp_file_path)
                    images = list(self._extract_images_xls(temp_file_path))

                logger.info(
                    "[excel_extractor] extraction finished",
                    extra={
                        "text_length": len(text_content or ""),
                        "image_count": len(images),
                    },
                )

                yield self.create_text_message(
                    text_content
                    if text_content.strip()
                    else "No textual content found in the Excel workbook."
                )

                for idx, (img_bytes, extension) in enumerate(images, start=1):
                    base_filename = original_filename.rsplit(".", 1)[0]
                    output_filename = f"{base_filename}_image_{idx}{extension}"
                    mime_type = self._resolve_mime_type(extension)

                    logger.info(
                        "[excel_extractor] yield image",
                        extra={
                            "index": idx,
                            "extension": extension,
                            "mime_type": mime_type,
                            "size": len(img_bytes),
                        },
                    )
                    yield self.create_blob_message(
                        blob=img_bytes,
                        meta={
                            "mime_type": mime_type,
                            "file_name": output_filename,
                        },
                    )
            finally:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
        except ValueError as exc:
            raise ValueError(f"Invalid value encountered: {exc}") from exc
        except Exception as exc:  # noqa: BLE001 - surface friendly error
            raise Exception(f"Error extracting from Excel workbook: {exc}") from exc

    def get_runtime_parameters(
        self,
        conversation_id: Optional[str] = None,
        app_id: Optional[str] = None,
        message_id: Optional[str] = None,
    ) -> list[ToolParameter]:
        return [
            ToolParameter(
                name="excel_content",
                label=I18nObject(en_US="Excel Content", zh_Hans="Excel 内容"),
                human_description=I18nObject(
                    en_US="Excel file (.xlsx/.xls) to extract text and embedded images from",
                    zh_Hans="要提取文本和图片的 Excel 文件(.xlsx/.xls)",
                ),
                type=ToolParameter.ToolParameterType.FILE,
                form=ToolParameter.ToolParameterForm.FORM,
                required=True,
                file_accepts=[
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                ],
            ),
        ]

    def _extract_text_xlsx(self, file_path: str) -> str:
        workbook = load_workbook(filename=file_path, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# Sheet: {sheet.title}")
            has_content = False
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                formatted_cells = [
                    self._format_cell_value(cell) for cell in row if cell not in (None, "")
                ]
                if not formatted_cells:
                    continue
                has_content = True
                row_text = " | ".join(formatted_cells)
                lines.append(f"Row {row_idx}: {row_text}")
            if not has_content:
                lines.append("No textual content in this sheet.")
            lines.append("")  # blank line between sheets
        return "\n".join(lines).strip()

    def _is_ooxml_excel(self, file_path: str) -> bool:
        if not is_zipfile(file_path):
            return False
        try:
            with ZipFile(file_path) as zipped_file:
                zipped_file.getinfo("[Content_Types].xml")
                zipped_file.getinfo("xl/workbook.xml")
                return True
        except Exception:
            return False

    def _extract_text_xls(self, file_path: str) -> str:
        workbook = xlrd.open_workbook(file_path, formatting_info=False)
        lines: list[str] = []
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            lines.append(f"# Sheet: {sheet_name}")
            has_content = False
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                formatted_cells = [
                    self._format_cell_value(cell) for cell in row if cell not in ("", None)
                ]
                if not formatted_cells:
                    continue
                has_content = True
                row_text = " | ".join(formatted_cells)
                lines.append(f"Row {row_idx + 1}: {row_text}")
            if not has_content:
                lines.append("No textual content in this sheet.")
            lines.append("")
        return "\n".join(lines).strip()

    def _extract_images_xlsx(self, file_path: str) -> Generator[tuple[bytes, str], None, None]:
        with ZipFile(file_path) as zipped_file:
            for member in zipped_file.namelist():
                if not member.startswith("xl/media/"):
                    continue
                extension = os.path.splitext(member)[1].lower() or ".png"
                with zipped_file.open(member) as image_file:
                    yield image_file.read(), extension

    def _extract_images_xls(self, file_path: str) -> Generator[tuple[bytes, str], None, None]:
        yielded = False
        for image in self._extract_images_from_ole_streams(file_path):
            yielded = True
            yield image

        if yielded:
            return

        for image in self._extract_images_from_bstore(file_path):
            yielded = True
            yield image

        if yielded:
            return

        data = Path(file_path).read_bytes()
        seen_ranges: list[tuple[int, int]] = []
        for signature, extension, handler in self._signature_handlers():
            search_start = 0
            while True:
                idx = data.find(signature, search_start)
                if idx == -1:
                    break

                if any(start <= idx < end for start, end in seen_ranges):
                    search_start = idx + 1
                    continue

                image_bytes = handler(data, idx)
                if not image_bytes:
                    search_start = idx + 1
                    continue

                end_idx = idx + len(image_bytes)
                seen_ranges.append((idx, end_idx))
                yield image_bytes, extension
                search_start = end_idx

    def _extract_images_from_ole_streams(
        self, file_path: str
    ) -> Generator[tuple[bytes, str], None, None]:
        if not olefile.isOleFile(file_path):
            return

        ole = olefile.OleFileIO(file_path)
        try:
            for entry in ole.listdir(streams=True):
                if not entry:
                    continue
                stream_name = entry[0].upper()
                if not stream_name.startswith("MBD"):
                    continue
                stream = ole.openstream(entry)
                try:
                    data = stream.read()
                finally:
                    stream.close()
                image = self._detect_image_from_stream(data)
                if image:
                    yield image
        finally:
            ole.close()

    def _extract_images_from_bstore(
        self, file_path: str
    ) -> Generator[tuple[bytes, str], None, None]:
        if not olefile.isOleFile(file_path):
            return

        ole = olefile.OleFileIO(file_path)
        try:
            if not ole.exists("Workbook"):
                return
            workbook_data = ole.openstream("Workbook").read()
        finally:
            ole.close()

        for record_type, payload in self._iter_biff_records_with_continue(workbook_data):
            if record_type != self._MSODRAWINGGROUP_RECORD:
                continue
            for image in self._extract_images_from_escher(payload):
                yield image
            break

    def _iter_biff_records_with_continue(
        self, data: bytes
    ) -> Generator[tuple[int, bytes], None, None]:
        position = 0
        total_length = len(data)
        while position + 4 <= total_length:
            record_type, length = struct.unpack_from("<HH", data, position)
            position += 4
            payload = data[position : position + length]
            position += length
            while position + 4 <= total_length:
                next_type, next_length = struct.unpack_from("<HH", data, position)
                if next_type != self._CONTINUE_RECORD:
                    break
                position += 4
                payload += data[position : position + next_length]
                position += next_length
            yield record_type, payload

    def _extract_images_from_escher(
        self, data: bytes
    ) -> Generator[tuple[bytes, str], None, None]:
        for _, _, record_type, record_data in self._iter_office_art_records(data):
            if record_type != self._BSE_RECORD_TYPE:
                continue
            decoded = self._decode_bse_image(record_data)
            if decoded:
                yield decoded

    def _iter_office_art_records(
        self, data: bytes
    ) -> Generator[tuple[int, int, int, bytes], None, None]:
        position = 0
        limit = len(data)
        while position + 8 <= limit:
            ver_instance, record_type, record_length = struct.unpack_from(
                "<HHI", data, position
            )
            record_ver = ver_instance & 0xF
            record_instance = ver_instance >> 4
            start = position + 8
            end = start + record_length
            if end > limit:
                break
            payload = data[start:end]
            yield record_ver, record_instance, record_type, payload
            if record_ver == 0xF:
                yield from self._iter_office_art_records(payload)
            position = end

    def _decode_bse_image(self, record_data: bytes) -> Optional[tuple[bytes, str]]:
        if not record_data:
            return None
        header_offset = self._locate_blip_record(record_data)
        if header_offset is None:
            return None
        if header_offset + 8 > len(record_data):
            return None
        _, record_type, record_length = struct.unpack_from(
            "<HHI", record_data, header_offset
        )
        start = header_offset + 8
        end = start + record_length
        if end > len(record_data):
            return None
        image_bytes = record_data[start:end]
        extension = self._resolve_bstore_extension(record_data[0], record_type)
        image_bytes = self._strip_blip_metadata(image_bytes, extension)
        if extension == ".bmp" and not image_bytes.startswith(b"BM"):
            image_bytes, extension = self._convert_dib_to_bmp(image_bytes)
        return image_bytes, extension

    def _locate_blip_record(self, record_data: bytes) -> Optional[int]:
        search_limit = min(len(record_data), 512)
        for offset in range(search_limit):
            if offset + 8 > len(record_data):
                break
            _, record_type, record_length = struct.unpack_from(
                "<HHI", record_data, offset
            )
            if record_type not in self._BLIP_RECORD_TYPES:
                continue
            if offset + 8 + record_length > len(record_data):
                continue
            return offset
        return None

    def _resolve_bstore_extension(self, blip_type: int, record_type: int) -> str:
        mapping = {
            0x02: ".emf",
            0x03: ".wmf",
            0x04: ".pict",
            0x05: ".jpg",
            0x06: ".png",
            0x07: ".bmp",
            0x11: ".tiff",
            0x12: ".jpg",
            0x13: ".jpg",
        }
        if blip_type in mapping:
            return mapping[blip_type]
        fallback = {
            0xF01D: ".jpg",
            0xF01E: ".png",
            0xF01F: ".bmp",
            0xF020: ".tiff",
            0xF021: ".jpg",
            0xF029: ".tiff",
            0xF02A: ".jpg",
            0xF02B: ".jpg",
            0xF02C: ".png",
        }
        return fallback.get(record_type, ".bin")

    def _strip_blip_metadata(self, image_bytes: bytes, extension: str) -> bytes:
        signatures: dict[str, tuple[bytes, ...]] = {
            ".jpg": (b"\xff\xd8\xff",),
            ".jpeg": (b"\xff\xd8\xff",),
            ".png": (b"\x89PNG\r\n\x1a\n",),
            ".gif": (b"GIF87a", b"GIF89a"),
            ".tif": (b"II*\x00", b"MM\x00*"),
            ".tiff": (b"II*\x00", b"MM\x00*"),
        }
        candidates = signatures.get(extension)
        if not candidates:
            return image_bytes
        for signature in candidates:
            idx = image_bytes.find(signature)
            if idx != -1:
                return image_bytes[idx:]
        return image_bytes

    def _convert_dib_to_bmp(self, dib_data: bytes) -> tuple[bytes, str]:
        if len(dib_data) < 40:
            return dib_data, ".dib"
        header_size = int.from_bytes(dib_data[0:4], "little")
        if header_size < 40 or header_size > len(dib_data):
            return dib_data, ".dib"
        bit_count = int.from_bytes(dib_data[14:16], "little")
        compression = int.from_bytes(dib_data[16:20], "little")
        colors_used = int.from_bytes(dib_data[32:36], "little")
        color_table_size = 0
        if bit_count in (1, 4, 8):
            entries = colors_used or (1 << bit_count)
            color_table_size = entries * 4
        elif bit_count in (16, 32) and compression == 3:
            color_table_size = 12
        if header_size + color_table_size > len(dib_data):
            color_table_size = max(0, len(dib_data) - header_size)
        pixel_offset = 14 + header_size + color_table_size
        file_size = 14 + len(dib_data)
        if pixel_offset > file_size:
            pixel_offset = 14 + header_size
        file_header = bytearray(14)
        file_header[0:2] = b"BM"
        file_header[2:6] = file_size.to_bytes(4, "little")
        file_header[10:14] = pixel_offset.to_bytes(4, "little")
        return bytes(file_header) + dib_data, ".bmp"

    def _signature_handlers(self):
        return (
            (b"\x89PNG\r\n\x1a\n", ".png", self._slice_png),
            (b"\xff\xd8\xff", ".jpg", self._slice_jpeg),
            (b"GIF87a", ".gif", self._slice_gif),
            (b"GIF89a", ".gif", self._slice_gif),
            (b"BM", ".bmp", self._slice_bmp),
        )

    def _detect_image_from_stream(self, data: bytes) -> Optional[tuple[bytes, str]]:
        for signature, extension, handler in self._signature_handlers():
            idx = data.find(signature)
            if idx != -1:
                return handler(data, idx), extension
        return None

    def _slice_png(self, data: bytes, start_idx: int) -> bytes:
        end_marker = b"IEND\xaeB`\x82"
        end_idx = data.find(end_marker, start_idx)
        if end_idx != -1:
            end_idx += len(end_marker)
        else:
            end_idx = len(data)
        return data[start_idx:end_idx]

    def _slice_jpeg(self, data: bytes, start_idx: int) -> bytes:
        end_marker = b"\xff\xd9"
        end_idx = data.find(end_marker, start_idx)
        if end_idx != -1:
            end_idx += len(end_marker)
        else:
            end_idx = len(data)
        return data[start_idx:end_idx]

    def _slice_gif(self, data: bytes, start_idx: int) -> bytes:
        trailer = b"\x3B"
        end_idx = data.find(trailer, start_idx)
        if end_idx != -1:
            end_idx += len(trailer)
        else:
            end_idx = len(data)
        return data[start_idx:end_idx]

    def _slice_bmp(self, data: bytes, start_idx: int) -> bytes:
        if start_idx + 6 <= len(data):
            size = int.from_bytes(data[start_idx + 2 : start_idx + 6], "little")
            if size > 0 and start_idx + size <= len(data):
                return data[start_idx : start_idx + size]
        return data[start_idx:]

    def _determine_suffix(self, filename: Optional[str]) -> str:
        if not filename:
            return ".xlsx"
        extension = os.path.splitext(filename)[1].lower()
        if extension in {".xlsx", ".xls"}:
            return extension
        return ".xlsx"

    def _format_cell_value(self, value: Any) -> str:
        if isinstance(value, float):
            value = value if value % 1 else int(value)
        return str(value)

    def _resolve_mime_type(self, extension: str) -> str:
        mime_types = {
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".gif": "image/gif",
            ".bmp": "image/bmp",
            ".tif": "image/tiff",
            ".tiff": "image/tiff",
            ".emf": "image/emf",
            ".wmf": "image/wmf",
            ".pict": "image/x-pict",
            ".dib": "image/bmp",
        }
        return mime_types.get(extension, "application/octet-stream")
